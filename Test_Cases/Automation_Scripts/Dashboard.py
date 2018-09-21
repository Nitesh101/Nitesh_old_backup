#!/usr/bin/python
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import os

List1=["S.No","TestCase ID","Expected Result","Actual Result","Status"]
def Heading():	
	font_color=Font(color=colors.BLUE)
	for count, statN in enumerate(List1):
		ws1.cell(row=1, column=count+1).font = font_color
		ws1.cell(row=1, column=count+1).value = statN

def dashboard(row1,col,mylist):
	global wb,ws1
	wb = load_workbook('dashboard.xlsx')
	ws1 = wb.active
	ws1 = wb.get_sheet_by_name("Test_Cases")
	Heading()
	max1 = ws1.max_row
	ws1.cell(row=row1+max1,column=col-1).value=row1-2 + max1       #for serial Number    
	ws1.cell(row=row1+max1,column=col).value=mylist[0]             #for TestCase ID
	ws1.cell(row=row1+max1,column=col+1).value=mylist[1]           #for actual  result
	ws1.cell(row=row1+max1,column=col+2).value=mylist[2]           #for expected status
	if(mylist[1]==mylist[2]):
		status="Pass"
		font_colors=Font(color=colors.GREEN)
	
	else:
		status="Fail"
		font_colors=Font(color=colors.RED)

	ws1.cell(row=row1+max1,column=col+3).font=font_colors	
	ws1.cell(row=row1+max1,column=col+3).value=status             #for status
	row1+=1

	wb.save('dashboard.xlsx')
	wb.close()

def create_dashboard(mylist):
	row1,col=2,2
	if os.path.exists("dashboard.xlsx"):
			dashboard(row1,col,mylist)
	else:	
		wb = Workbook()
		ws1 = wb.active
		ws1.title = "Test_Cases"
		#ws2 = wb.create_sheet(title="Report_Page")
		wb.save("dashboard.xlsx")
		dashboard(row1,col,mylist)
